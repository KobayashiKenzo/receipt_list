import pymupdf as fitz
import pytesseract
import textdistance
import pandas as pd
import jellyfish
import re
import os
from PIL import Image, ImageEnhance, ImageFilter
from pdf2image import convert_from_path
from difflib import get_close_matches
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict

# perfects!
class MedicalReceiptProcessor:
    def __init__(self, template_path: str, phonebook_path: str):
        self.template = template_path
        self.start_row = 9
        # phonebook.csvの読み込み（文字列型として指定）
        self.phonebook = pd.read_csv(phonebook_path, dtype={'area_code': str, 'phone_number1': str, 'phone_number2': str})
        
        # 桁数を揃える処理を追加
        self.phonebook['area_code'] = self.phonebook['area_code'].apply(lambda x: x.zfill(2))
        self.phonebook['phone_number1'] = self.phonebook['phone_number1'].apply(lambda x: x.zfill(4))
        self.phonebook['phone_number2'] = self.phonebook['phone_number2'].apply(lambda x: x.zfill(4))

    def correct_patient_name(self, ocr_name: str, method="levenshtein") -> str:
        """OCR結果から最も近い患者名を選択"""
        if not ocr_name or ocr_name == "不明":
            return "不明"
            
        # 患者データの読み込み
        patient_df = pd.read_csv('patients.csv', encoding='utf-8-sig')
        
        # OCR結果の前処理
        ocr_name = re.sub(r'[^\w\s]', '', ocr_name).strip()  # 記号を削除し、余分な空白をトリム
        
        # 姓または名の部分一致を確認
        for _, row in patient_df.iterrows():
            family = row['family_name']
            first = row['first_name']
            full_name = family + first
            
            if ocr_name == family or ocr_name == first or ocr_name == full_name:
                return full_name
        
        # 完全な名前のリストを作成
        full_names = [(row['family_name'] + row['first_name']) for _, row in patient_df.iterrows()]
        
        if method == "levenshtein":
            for cutoff in [0.6, 0.5, 0.4]:  # 段階的にcutoff値を下げる
                closest_match = get_close_matches(ocr_name, full_names, n=1, cutoff=cutoff)
                if closest_match:
                    return closest_match[0]
            return "不明"
        elif method == "jaro_winkler":
            scores = {name: textdistance.jaro_winkler(ocr_name, name) for name in full_names}
            closest_match = max(scores, key=scores.get)
            return closest_match if scores[closest_match] > 0.6 else "不明"
        
        return "不明"

    def get_facility_name_by_phone(self, phone_number: str) -> str:
        """電話番号から施設名を取得"""
        if phone_number == "不明":
            return "不明"
        
        # ハイフンを削除
        clean_number = phone_number.replace('-', '')
        
        # 電話番号のパターンに応じて分割
        if len(clean_number) >= 10:
            for i in range(2, 5):
                if i <= len(clean_number):
                    area_code = clean_number[:i]
                    remaining = clean_number[i:]
                    
                    if len(remaining) >= 4:
                        phone_part1 = remaining[:4]
                        
                        # 文字列として比較
                        area_code_str = str(area_code)
                        phone_part1_str = str(phone_part1)
                        
                        # デバッグ出力
                        print(f"検索: area_code={area_code_str}, phone_part1={phone_part1_str}")
                        
                        # エリアコードとphone_number1が一致する場合
                        match1 = self.phonebook[
                            (self.phonebook['area_code'].astype(str) == area_code_str) & 
                            (self.phonebook['phone_number1'].astype(str) == phone_part1_str)
                        ]
                        
                        if not match1.empty:
                            return match1.iloc[0]['facility_name']

        return "不明"


    def preprocess_image(self, pdf_path: str) -> str:
        """PDFまたは画像ファイルを前処理してOCR精度を向上させる"""
        try:
            # 言語データの存在確認
            tessdata_dir = os.environ.get("TESSDATA_PREFIX", "/usr/share/tesseract-ocr/4.00/tessdata")
            if not os.path.exists(os.path.join(tessdata_dir, 'jpn.traineddata')):
                raise Exception(f"日本語の言語データが見つかりません: {tessdata_dir}")

            # PDFの場合、画像に変換
            if pdf_path.endswith(".pdf"):
                images = convert_from_path(pdf_path)
            else:
                images = [Image.open(pdf_path)]

            text_results = []
            for image in images:
                # 解像度の向上
                image = image.resize((image.width * 3, image.height * 3), Image.LANCZOS)
                # グレースケール変換
                image = image.convert("L")
                # コントラスト強調
                enhancer = ImageEnhance.Contrast(image)
                image = enhancer.enhance(1.5)
                # シャープネス強調
                # image = image.filter(ImageFilter.SHARPEN)
                # 二値化
                threshold = 220
                image = image.point(lambda x: 0 if x < threshold else 255, '1')
                # OCR処理（エラーハンドリング付き）
                try:
                    config = "--psm 3 --oem 1"
                    text = pytesseract.image_to_string(image, lang="jpn", config=config)
                    text_results.append(text)
                except pytesseract.TesseractError as e:
                    print(f"OCR処理エラー: {str(e)}")
                    raise

            return "\n".join(text_results)

        except Exception as e:
            raise Exception(f"PDFまたは画像処理エラー: {str(e)}")

    def clean_ocr_text(self, text):
        """OCR結果から不可視文字・全角スペースを削除し、令和の表記も正規化"""
        # 不可視文字の削除
        text = text.replace("\u200b", "").replace("\xa0", "").replace("\ufeff", "")        
        # 令和の年号のゼロ埋めを削除（例: 令和06年 → 令和6年）
        text = re.sub(r'令和0?(\d+)', r'令和\1', text)
        # 連続スペースを1つに統一
        text = re.sub(r'[\s　]+', ' ', text)

        return text.strip()

    def extract_data(self, pdf_path: str) -> pd.DataFrame:
        """PDFからデータを抽出し、OCR結果を統合"""
        try:
            # synOCRテキスト化
            doc = fitz.open(pdf_path)
            synocr_text = "\n".join([page.get_text() for page in doc])
            synocr_text = self.clean_ocr_text(synocr_text)        
            # print("=== synOCR抽出テキスト ===")
            # print(synocr_text)
            # synocr_numbers = re.findall(pattern_number, synocr_text)
            # print(f'synocr番号のみ:\n{synocr_numbers}')
            # Tesseractテキスト化
            tess_text = self.preprocess_image(pdf_path)
            tess_text = self.clean_ocr_text(tess_text)
            # 連続する漢字の間の半角スペースを削除        
            # print("=== Tesseract抽出テキスト ===")
            # print(tess_text)
            # tess_numbers = re.findall(pattern_number, tess_text)
            # print(f'tess番号のみ:\n{tess_numbers}')

            # 施設名
            # phone_patterns = [
            #     r'\b(0\d)\W*(\d{4})\W*(\d{4})\b',
            #     r'(?:TEL|電話|T)?\s*[:：]?\s*(\d{2,4})[-−](\d{2,4})[-−](\d{4})',
            #     r'(\d{2})[-−](\d{4})[-−](\d{4})',
            #     r'(?<!\d)(03-\d{4}-\d{4})(?!\d)',
            #     r'.*?(\d{2})[-−](\d{4})[-−](\d{4})'  # より柔軟なパターン
            # ]

            phone_patterns = [r'(?:TEL|電話|T)?\s*[:：]?\s*(0\d)[\s\-–—]*?(\d{4})[\s\-–—]*?(\d{4})\b']
            phone_number = "不明"
            # まずsynOCRテキストで検索
            for pattern in phone_patterns:
                phone_match = re.search(pattern, synocr_text)
                if phone_match:
                    area_code = phone_match.group(1)
                    phone_part1 = phone_match.group(2)
                    phone_part2 = phone_match.group(3)
                    phone_number = f"{area_code}-{phone_part1}-{phone_part2}"
                    # print(f"抽出された電話番号(synOCR): {phone_number}")
                    break

            # synOCRで見つからなかった場合、Tesseractテキストも確認
            if phone_number == "不明":
                for pattern in phone_patterns:
                    phone_match = re.search(pattern, tess_text)
                    if phone_match:
                        area_code = phone_match.group(1)
                        phone_part1 = phone_match.group(2)
                        phone_part2 = phone_match.group(3)
                        phone_number = f"{area_code}-{phone_part1}-{phone_part2}"
                        # print(f"抽出された電話番号(Tesseract): {phone_number}")
                        break

            # print(f"phonebook.csvの内容: {self.phonebook}")
            facility_name = self.get_facility_name_by_phone(phone_number)

            final_result = {}

            # synOCR, Tesseractそれぞれからデータを抽出
            synocr_result = self._parse_text(synocr_text, facility_name)
            tess_result = self._parse_text(tess_text, facility_name)

            # デフォルト値を設定
            amount_from_synocr = synocr_result.get('amount', "0") if synocr_result else "0"
            amount_from_tess = tess_result.get('amount', "0") if tess_result else "0"

            # カンマを削除
            amount_from_synocr = re.sub(r'[^\d]', '', amount_from_synocr.replace(',', '').strip())
            amount_from_tess = re.sub(r'[^\d]', '', amount_from_tess.replace(',', '').strip())
            synocr_amount = int(amount_from_synocr) if amount_from_synocr.isdigit() else 0
            tess_amount = int(amount_from_tess) if amount_from_tess.isdigit() else 0

            if synocr_amount > 0:
                final_result['amount'] = synocr_amount
            elif tess_amount > 0:
                final_result['amount'] = tess_amount
            elif synocr_amount == 0:
                final_result['amount'] = synocr_amount
            elif tess_amount == 0:
                final_result['amount'] = tess_amount
            else:
                final_result['amount'] = 'amount取得Error'

            
            # 患者名補正
            patient_df = pd.read_csv('patients.csv', encoding='utf-8-sig')

            # 新しい方法：family_nameとfirst_nameを組み合わせて完全な名前のリストを作成
            # patient_list = [(row['family_name'] + row['first_name']) for _, row in patient_df.iterrows()]
            synocr_patient = synocr_result.get('patient', "不明")
            tess_patient = tess_result.get('patient', "不明")
            corrected_synocr = "不明"
            corrected_tess = "不明"

            # 両方の結果が有効なら、どちらがリストとより近いか評価する
            if synocr_patient != "不明" and tess_patient != "不明":
                corrected_synocr = self.correct_patient_name(synocr_patient, method="levenshtein")
                corrected_tess = self.correct_patient_name(tess_patient, method="levenshtein")
                # Jaro-Winklerスコアで評価して選択
                synocr_score = textdistance.jaro_winkler(synocr_patient, corrected_synocr)
                tess_score = textdistance.jaro_winkler(tess_patient, corrected_tess)
                chosen_patient = corrected_synocr if synocr_score >= tess_score else corrected_tess
            elif synocr_patient != "不明":
                corrected_synocr = self.correct_patient_name(synocr_patient, method="levenshtein")
                chosen_patient = corrected_synocr if corrected_synocr != "不明" else synocr_patient
            elif tess_patient != "不明":
                corrected_tess = self.correct_patient_name(tess_patient, method="levenshtein")
                chosen_patient = corrected_tess if corrected_tess != "不明" else "不明"
            else:
                chosen_patient = "不明"


            final_result['patient'] = chosen_patient

            # 日付処理
            synocr_date = synocr_result.get('date', "不明")
            tess_date = tess_result.get('date', "不明")

            if synocr_date != "不明" and tess_date != "不明":
                # 両方の日付が存在する場合、より新しい方を採用
                final_result['date'] = max(synocr_date, tess_date)
            elif synocr_date != "不明":
                final_result['date'] = synocr_date
            elif tess_date != "不明":
                final_result['date'] = tess_date
            else:
                final_result['date'] = "不明"

            # 施設名
            final_result['institution'] = facility_name
            print(f'institution名抽出{facility_name}')
            is_pharmacy = '薬局' in final_result.get('institution', '')
            final_result['categories'] = ['医薬品購入' if is_pharmacy else '診療・治療']

            print(f"最終結果: {final_result}\n")  # デバッグ用ログ

            return pd.DataFrame([final_result])

        except Exception as e:
            print(f"データ抽出エラー: {str(e)}")
            
            return pd.DataFrame(columns=['patient', 'institution', 'date', 'amount', 'categories'])

    def post_process_text(self, text: str) -> str:

        # 数字の正規化（カンマや円の除去など）
        text = re.sub(r'([0-9]),([0-9])', r'\1\2', text)
        text = text.replace('円', '')

        # 全角数字を半角に変換
        zen = "０１２３４５６７８９"
        han = "0123456789"
        trans_table = str.maketrans(zen, han)
        text = text.translate(trans_table)
        
        # 日本語の文字化けを修正
        text = text.replace('ー', '-')
        text = text.replace('（', '(')
        text = text.replace('）', ')')

        return text.strip()

    def _parse_text(self, text: str, facility_name: str) -> dict:
        """テキストから医療費データを解析"""

        # 金額キーワード辞書を読み込む
        try:
            words_df = pd.read_csv('amount_words.csv', encoding='utf-8-sig')

            # 施設名がカラムに存在する場合、そのカラムを参照
            if facility_name in words_df.columns:
                amount_keywords = words_df[facility_name].dropna().tolist()
            else:
                raise ValueError(f"施設名 '{facility_name}' が辞書に存在しません")
        except Exception as e:
            print(f"金額キーワード辞書の読み込みエラー: {str(e)}")
            # エラー時にはデフォルトキーワードを返す
            amount_keywords = ["請求額", "領収額"]

        # デバッグ: 読み込んだキーワードリストを確認
        print("参照したキーワードリスト:", amount_keywords)
        # キーワード～円(＋ハイフン)の間の文字をキャプチャして、数字以外を削除
        for keyword in amount_keywords:
            escaped_keyword = re.escape(keyword)
            # 「円」の後ろに続くハイフン類もキャプチャ対象にする
            pattern = rf'({escaped_keyword})(.*?)(円[ー-]*)'

            def keep_digits_only(m):
                # m.group(1) => キーワード
                # m.group(2) => キーワードと「円」(+ハイフン)の間の文字
                # m.group(3) => "円ー..." の部分
                # 数字以外の文字を削除
                digits_only = re.sub(r'\D+', '', m.group(2))
                # 最終的には "円" だけ残してハイフンは削除
                return f"{m.group(1)}{digits_only}円"

            text = re.sub(pattern, keep_digits_only, text)

        # ここで金額抽出用の正規表現パターンを作成
        amount_patterns = [
            rf'{kw}\s*(\d+)\s*円'
            for kw in amount_keywords
        ]
        amount_patterns.append(r'(\d{1,3}(?:,\d{3})*)\s*円')  # 汎用パターン

        # 共通の正規表現（患者名や日付など）
        patterns = {
            'amount': amount_patterns,
            'patient': [
                r'(小[^\n]{1,8}\s*三|[^\n]{2,10}\s*(様|賢三))'
            ],
            'date': [
                r'令\s*和\s*(\d{1,2})[-/\s年]\s*(\d{1,2})[-/\s月]\s*(\d{1,2})\s*[日]?',
            ]
        }

        result = {}
        for key, patterns_list in patterns.items():
            for pattern in patterns_list:
                match = re.search(pattern, text)
                if match:
                    if key == 'amount':
                        result[key] = re.sub(r'[^\d]', '', match.group(1))  # 数字以外を削除
                    elif key == 'patient':
                        patient_name = match.group(1).replace('様', '').strip()
                        result[key] = patient_name
                    elif key == 'date':
                        year = int(match.group(1)) + 2018
                        month = match.group(2).zfill(2)
                        day = match.group(3).zfill(2)
                        result[key] = f"{year}-{month}-{day}"
                    break

            if key not in result:
                result[key] = "不明"

        print(f"抽出結果: {result}")  # デバッグ用ログ
        return result

    def update_excel(self, data: pd.DataFrame, output_file: str) -> bool:
        """Excelファイルにデータを書き込む"""
        try:
            if data.empty:
                raise ValueError("データが空です")

            file_path = Path(output_file)
            wb = load_workbook(output_file) if file_path.exists() else load_workbook(self.template)
            ws = wb.active

            last_row = max((cell.row for cell in ws['C'] if cell.row >= self.start_row and cell.value), default=self.start_row - 1)
            next_empty_row = last_row + 1

            for idx, row in data.iterrows():
                target_row = next_empty_row + idx
                ws[f"B{target_row}"] = row.get('patient', "不明")
                ws[f"C{target_row}"] = row.get('institution', "不明")
                ws[f"H{target_row}"] = row.get('amount', 0)
                ws[f"J{target_row}"] = row.get('date', "不明")
                ws[f"E{target_row}" if '薬局' in row.get('institution', '') else f"D{target_row}"] = "該当する"

            wb.save(output_file)
            print(f"Excelファイル {output_file} に保存しました")
            return True

        except Exception as e:
            print(f"Excel更新エラー: {str(e)}")
            return False

